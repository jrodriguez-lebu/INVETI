#!/usr/bin/env python3
"""
Import script: Inventario informatica 2025.xlsx -> Laravel SQLite database.
Reads the 'PC' sheet, cleans data, and inserts into:
  - departamentos
  - funcionarios
  - equipos
Existing sample data in those three tables is cleared first.
tipos_equipo and actas are left untouched.
"""

import sqlite3
import json
import os
import sys
import re
from openpyxl import load_workbook

# ── Paths ────────────────────────────────────────────────────────────────────
EXCEL_PATH = "/Users/macbook/Downloads/Inventario informática 2025.xlsx"
DB_PATH    = "/Users/macbook/Herd/INVETI/database/database.sqlite"

# ── TIPO EQUIPO → tipos_equipo.nombre mapping ────────────────────────────────
TIPO_MAP = {
    "AIO":          "AIO",
    "NOTEBOOK":     "Notebook",
    "NOTEBBOK":     "Notebook",
    "IMPRESORA":    "Impresora",
    "MULTIFUNCIONAL": "Multifuncional",
    "MONITOR":      "Monitor",
    "TABLET":       "Tablet",
    "IPAD":         "Tablet",
    "ESCANNER":     "Otro",
    "NVR":          "Otro",
    "NO TIENE":     "Otro",
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def clean(val):
    """Strip whitespace; return None for empty/None values."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def normalize_tipo(raw):
    """Normalise a raw TIPO EQUIPO string using TIPO_MAP."""
    if raw is None:
        return None
    return TIPO_MAP.get(raw.strip().upper(), None)


def parse_fecha(val):
    """
    Accept integers (year only) or strings like 'SEP-2025', 'SEP 2025',
    '2025', '01/01/2025' and return a YYYY-MM-DD string or None.
    """
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None

    # Pure year (e.g. 2025 or '2025')
    if re.fullmatch(r'\d{4}', s):
        return s + "-01-01"

    # Float that looks like a year (e.g. 2025.0)
    try:
        f = float(s)
        if 1990 <= f <= 2100:
            return str(int(f)) + "-01-01"
    except ValueError:
        pass

    # Month-Year strings like SEP-2025 or SEP 2025
    month_map = {
        "ENE": "01", "FEB": "02", "MAR": "03", "ABR": "04",
        "MAY": "05", "JUN": "06", "JUL": "07", "AGO": "08",
        "SEP": "09", "OCT": "10", "NOV": "11", "DIC": "12",
        "JAN": "01", "APR": "04", "AUG": "08", "DEC": "12",
    }
    m = re.match(r'^([A-Za-z]{3})[-\s](\d{4})$', s)
    if m:
        mon = m.group(1).upper()
        yr  = m.group(2)
        num = month_map.get(mon)
        if num:
            return f"{yr}-{num}-01"

    # DD/MM/YYYY or YYYY-MM-DD
    m = re.match(r'^(\d{2})/(\d{2})/(\d{4})$', s)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"

    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})$', s)
    if m:
        return s

    # Last resort: return None
    return None


def convert_codigo(val):
    """
    Convert a float inventory code (e.g. 10701001.12) to string '10701001-12'.
    Non-float or already-string values are returned as-is (stripped).
    """
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    # If it looks like a float (digits.digits), replace dot with dash
    m = re.match(r'^(\d+)\.(\d+)$', s)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    return s


def split_name(full_name):
    """
    Split 'Primer Segundo Apellido' → (nombre='Primer', apellido='Segundo Apellido').
    Single word → (nombre=word, apellido='').
    """
    parts = full_name.strip().split()
    if len(parts) >= 2:
        return parts[0], " ".join(parts[1:])
    return parts[0], ""


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: Excel file not found at {EXCEL_PATH}", file=sys.stderr)
        sys.exit(1)

    if not os.path.exists(DB_PATH):
        print(f"ERROR: SQLite database not found at {DB_PATH}", file=sys.stderr)
        sys.exit(1)

    print("Loading Excel file…")
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["PC"]

    # Read header row
    headers = [clean(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"  Columns found: {headers}")

    # Build column index map (case-insensitive, strip whitespace)
    col_idx = {h.upper() if h else "": i for i, h in enumerate(headers)}

    def get(row, name):
        """Get cell value from row by column name (upper)."""
        idx = col_idx.get(name.upper())
        if idx is None:
            return None
        return clean(row[idx].value)

    # Collect all data rows — skip completely empty rows (Excel has formatting beyond data)
    all_rows = list(ws.iter_rows(min_row=2, values_only=False))
    rows = [r for r in all_rows if any(c.value is not None for c in r)]
    print(f"  Total data rows (non-empty): {len(rows)}")

    # ── Connect to SQLite ─────────────────────────────────────────────────────
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = OFF")
    cur = conn.cursor()

    # ── Clear existing sample data ────────────────────────────────────────────
    print("\nClearing existing data (equipos, funcionarios, departamentos)…")
    cur.execute("DELETE FROM equipos")
    cur.execute("DELETE FROM funcionarios")
    cur.execute("DELETE FROM departamentos")
    # Reset autoincrement counters
    cur.execute("DELETE FROM sqlite_sequence WHERE name IN ('equipos','funcionarios','departamentos')")
    conn.commit()
    print("  Done.")

    # ── Load tipos_equipo lookup ───────────────────────────────────────────────
    cur.execute("SELECT id, nombre FROM tipos_equipo")
    tipo_rows = cur.fetchall()
    tipo_lookup = {nombre.strip(): tid for tid, nombre in tipo_rows}
    print(f"\ntipos_equipo found: {tipo_lookup}")

    # ── Pass 1: Collect unique DIRECCION values → departamentos ───────────────
    print("\nCreating departamentos…")
    unique_dirs = set()
    for row in rows:
        d = get(row, "DIRECCION")
        if d:
            unique_dirs.add(d)

    dept_id_map = {}  # name → id
    now = "2026-04-11 00:00:00"
    for nombre in sorted(unique_dirs):
        cur.execute(
            "INSERT INTO departamentos (nombre, descripcion, created_at, updated_at) VALUES (?, ?, ?, ?)",
            (nombre, None, now, now)
        )
        dept_id_map[nombre] = cur.lastrowid

    conn.commit()
    print(f"  Created {len(dept_id_map)} departamentos.")

    # ── Pass 2: Collect unique ENCARGADO names → funcionarios ─────────────────
    print("\nCreating funcionarios…")
    # Key = normalised full-name uppercase; value = row data
    funcionario_map = {}  # full_name_upper → db id

    # First gather unique combos: (ENCARGADO, CARGO, DIRECCION)
    # We pick the first CARGO we see for a given ENCARGADO name
    encargado_data = {}  # name_upper → {nombre, apellido, cargo, departamento_id}
    for row in rows:
        encargado = get(row, "ENCARGADO")
        if not encargado:
            continue
        key = encargado.upper()
        if key in encargado_data:
            continue
        cargo     = get(row, "CARGO") or "Sin cargo"
        direccion = get(row, "DIRECCION")
        dept_id   = dept_id_map.get(direccion) if direccion else None

        nombre, apellido = split_name(encargado)
        encargado_data[key] = {
            "nombre":        nombre,
            "apellido":      apellido,
            "cargo":         cargo,
            "departamento_id": dept_id,
            "full_name":     encargado,
        }

    # Insert funcionarios (rut must be unique; generate a placeholder)
    rut_counter = 10000000
    for key, data in encargado_data.items():
        if data["departamento_id"] is None:
            # Need a departamento; create a placeholder "SIN DIRECCION" if needed
            if "SIN DIRECCION" not in dept_id_map:
                cur.execute(
                    "INSERT INTO departamentos (nombre, descripcion, created_at, updated_at) VALUES (?, ?, ?, ?)",
                    ("SIN DIRECCION", "Sin dirección asignada", now, now)
                )
                dept_id_map["SIN DIRECCION"] = cur.lastrowid
            data["departamento_id"] = dept_id_map["SIN DIRECCION"]

        rut_placeholder = f"00.000.{rut_counter:03d}-0"
        rut_counter += 1

        cur.execute(
            """INSERT INTO funcionarios
               (nombre, apellido, rut, cargo, departamento_id, email, telefono, activo, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                data["nombre"],
                data["apellido"],
                rut_placeholder,
                data["cargo"],
                data["departamento_id"],
                None,  # email
                None,  # telefono
                1,     # activo
                now,
                now,
            )
        )
        funcionario_map[key] = cur.lastrowid

    conn.commit()
    print(f"  Created {len(funcionario_map)} funcionarios.")

    # ── Pass 3: Insert equipos ─────────────────────────────────────────────────
    print("\nImporting equipos…")
    inserted   = 0
    skipped    = 0
    incomplete = 0
    inv_counter_map = {}  # track used numero_inventario to avoid duplicates
    inv_auto_seq = 1

    for row in rows:
        # Raw values
        direccion   = get(row, "DIRECCION")
        encargado   = get(row, "ENCARGADO")
        tipo_raw    = get(row, "TIPO EQUIPO")
        marca       = get(row, "MARCA")
        modelo      = get(row, "MODELO") or "Sin modelo"
        cpu         = get(row, "CPU")
        ram         = get(row, "RAM")
        almacenamiento = get(row, "ALMACENAMIENTO")
        ubicacion   = get(row, "Ubicación")
        anydesk     = get(row, "Código AnyDesk")
        office      = get(row, "Office16 / Office365")
        clave_win   = get(row, "Clave Windows 10Pro")
        comentario  = get(row, "Comentario")
        fecha_raw   = get(row, "FECHA ADQUI")
        factura     = get(row, "Factura, OC, u otro documento ")
        codigo_inv  = get(row, "Codigo inventario")

        # Derived values
        tipo_nombre = normalize_tipo(tipo_raw) if tipo_raw else None
        tipo_id     = tipo_lookup.get(tipo_nombre) if tipo_nombre else None
        dept_id     = dept_id_map.get(direccion) if direccion else None
        func_key    = encargado.upper() if encargado else None
        func_id     = funcionario_map.get(func_key) if func_key else None

        # Numero inventario
        numero_inv_raw = convert_codigo(codigo_inv)
        if numero_inv_raw:
            # Deduplicate: append suffix if collision
            base = numero_inv_raw
            suffix = 0
            while numero_inv_raw in inv_counter_map:
                suffix += 1
                numero_inv_raw = f"{base}-dup{suffix}"
            inv_counter_map[numero_inv_raw] = True
        else:
            numero_inv_raw = f"INV-{inv_auto_seq:04d}"
            while numero_inv_raw in inv_counter_map:
                inv_auto_seq += 1
                numero_inv_raw = f"INV-{inv_auto_seq:04d}"
            inv_counter_map[numero_inv_raw] = True
            inv_auto_seq += 1

        # Fecha adquisicion
        fecha_adq = parse_fecha(fecha_raw)

        # Incomplete detection
        campos_faltantes = []
        if not encargado:
            campos_faltantes.append("Encargado")
        if not tipo_raw:
            campos_faltantes.append("Tipo Equipo")
        if not marca:
            campos_faltantes.append("Marca")
        if not direccion:
            campos_faltantes.append("Dirección")

        datos_incompletos = 1 if campos_faltantes else 0
        if datos_incompletos:
            incomplete += 1

        # Default marca if missing (we still insert the record)
        marca_val = marca or "Sin marca"

        # tipo_id fallback: use 'Otro' if tipo_raw exists but not in map
        if tipo_raw and tipo_id is None:
            tipo_id = tipo_lookup.get("Otro")

        # If still no tipo_id, use 'Otro'
        if tipo_id is None:
            tipo_id = tipo_lookup.get("Otro")

        campos_json = json.dumps(campos_faltantes, ensure_ascii=False) if campos_faltantes else None

        try:
            cur.execute(
                """INSERT INTO equipos
                   (tipo_equipo_id, marca, modelo, numero_serie, numero_inventario,
                    estado, funcionario_id, departamento_id,
                    fecha_adquisicion, valor_adquisicion, descripcion, observaciones,
                    cpu, ram, almacenamiento, ubicacion, codigo_anydesk,
                    office_version, clave_windows, factura_documento,
                    datos_incompletos, campos_faltantes,
                    created_at, updated_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    tipo_id,
                    marca_val,
                    modelo,
                    None,               # numero_serie (not in Excel)
                    numero_inv_raw,
                    "activo",
                    func_id,
                    dept_id,
                    fecha_adq,
                    None,               # valor_adquisicion
                    comentario,         # descripcion
                    None,               # observaciones
                    cpu,
                    ram,
                    almacenamiento,
                    ubicacion,
                    anydesk,
                    office,
                    clave_win,
                    factura,
                    datos_incompletos,
                    campos_json,
                    now,
                    now,
                )
            )
            inserted += 1
        except sqlite3.IntegrityError as e:
            print(f"  SKIP row (IntegrityError): {e} | inv={numero_inv_raw}")
            skipped += 1

    conn.commit()
    conn.execute("PRAGMA foreign_keys = ON")
    conn.close()

    # ── Summary ───────────────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("IMPORT COMPLETE")
    print("=" * 60)
    print(f"  Departamentos created : {len(dept_id_map)}")
    print(f"  Funcionarios created  : {len(funcionario_map)}")
    print(f"  Equipos inserted      : {inserted}")
    print(f"  Equipos skipped       : {skipped}")
    print(f"  Equipos incompletos   : {incomplete} ({incomplete * 100 // inserted if inserted else 0}%)")
    print("=" * 60)


if __name__ == "__main__":
    main()
